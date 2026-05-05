"""
_build_excel.py — generates greek_banking_model.xlsx
Run: python deliverables/_build_excel.py
"""
import sqlite3, sys
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
DB   = ROOT / "02_Banking_Sector_Dashboard" / "data" / "greek_banking_final.db"
OUT  = Path(__file__).parent / "greek_banking_model.xlsx"

BANKS = ["Eurobank", "Alpha Bank", "Piraeus Bank", "NBG"]
YEARS = [2022, 2023, 2024]

BANK_COLORS = {
    "Eurobank":     "0067B1",
    "Alpha Bank":   "E2231A",
    "Piraeus Bank": "F7A600",
    "NBG":          "003087",
}

# ── Style constants ───────────────────────────────────────────────────────────
NAVY        = "0A1628"
DARK_BLUE   = "1B3A6B"
MID_BLUE    = "2D5A9E"
LIGHT_BLUE  = "D6E4F7"
LIGHTEST    = "EEF4FC"
WHITE       = "FFFFFF"
GREEN       = "1A7340"
LIGHT_GREEN = "D6F0E0"
AMBER       = "B45309"
LIGHT_AMBER = "FEF3C7"
RED         = "991B1B"
LIGHT_RED   = "FEE2E2"
GREY_BG     = "F1F5F9"
BORDER_COL  = "CBD5E1"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def border(style="thin"):
    s = Side(style=style, color=BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)

def thin_bottom():
    s = Side(style="thin", color=BORDER_COL)
    return Border(bottom=s)

def header_style(ws, row, col, text, bg=DARK_BLUE, fg=WHITE,
                 bold=True, size=10, center=True, border_style=True):
    c = ws.cell(row=row, column=col, value=text)
    c.fill    = fill(bg)
    c.font    = font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(horizontal="center" if center else "left",
                             vertical="center", wrap_text=True)
    if border_style:
        c.border = border()
    return c

def data_cell(ws, row, col, value, fmt=None, bg=WHITE, fg="0F172A",
              bold=False, center=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", size=10, bold=bold, color=fg)
    c.fill      = fill(bg)
    c.alignment = Alignment(horizontal="center" if center else "left",
                             vertical="center")
    c.border    = border()
    if fmt:
        c.number_format = fmt
    return c


# ── Load data ─────────────────────────────────────────────────────────────────
con    = sqlite3.connect(DB)
kpis   = pd.read_sql("SELECT * FROM kpis ORDER BY bank, year", con)
income = pd.read_sql("SELECT * FROM income_statement", con)
bs_raw = pd.read_sql("SELECT * FROM balance_sheet", con)
con.close()

income_p = income.pivot_table(index=["bank","year"], columns="metric", values="value").reset_index()
income_p.columns.name = None
bs_p     = bs_raw.pivot_table(index=["bank","year"], columns="metric", values="value").reset_index()
bs_p.columns.name = None

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER
# ═══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Cover"
ws.sheet_view.showGridLines = False

# Title block
ws.row_dimensions[1].height = 10
ws.row_dimensions[2].height = 50
ws.row_dimensions[3].height = 30
ws.row_dimensions[4].height = 20
ws.row_dimensions[5].height = 20
ws.row_dimensions[6].height = 20
ws.row_dimensions[7].height = 20
ws.row_dimensions[8].height = 10

ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 35
ws.column_dimensions["C"].width = 55
ws.column_dimensions["D"].width = 20

from openpyxl.styles import PatternFill
for row in range(1, 30):
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = fill(NAVY)

# Main title
c = ws.cell(row=2, column=2, value="GREEK BANKING SECTOR ANALYSIS")
c.font = Font(name="Calibri", bold=True, size=22, color=WHITE)
c.alignment = Alignment(horizontal="left", vertical="center")

c2 = ws.cell(row=3, column=2, value="Eurobank · Alpha Bank · Piraeus Bank · NBG  |  FY2022–2024")
c2.font = Font(name="Calibri", size=13, color="94A3B8")
c2.alignment = Alignment(horizontal="left", vertical="center")

# Table of contents
toc = [
    ("Cover",            "This sheet — overview and table of contents"),
    ("Sector Summary",   "Cross-bank KPI comparison and key metrics (2022–2024)"),
    ("Eurobank",         "Full IS, BS, and KPIs for Eurobank"),
    ("Alpha Bank",       "Full IS, BS, and KPIs for Alpha Bank"),
    ("Piraeus Bank",     "Full IS, BS, and KPIs for Piraeus Bank"),
    ("NBG",              "Full IS, BS, and KPIs for NBG"),
    ("Assumptions",      "Forecast assumptions — ECB rate scenarios, growth rates"),
]

ws.cell(row=9, column=2, value="TABLE OF CONTENTS").font = Font(name="Calibri", bold=True, size=12, color="60A5FA")

for i, (sheet, desc) in enumerate(toc, 11):
    ws.cell(row=i, column=2, value=sheet).font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    ws.cell(row=i, column=3, value=desc).font  = Font(name="Calibri", size=10, color="94A3B8")

ws.cell(row=20, column=2, value="Analyst:").font = Font(name="Calibri", bold=True, size=10, color="60A5FA")
ws.cell(row=20, column=3, value="Spyros Papastergiou").font = Font(name="Calibri", size=10, color=WHITE)
ws.cell(row=21, column=2, value="Contact:").font = Font(name="Calibri", bold=True, size=10, color="60A5FA")
ws.cell(row=21, column=3, value="spyrossyo96@gmail.com").font = Font(name="Calibri", size=10, color=WHITE)
ws.cell(row=22, column=2, value="Data source:").font = Font(name="Calibri", bold=True, size=10, color="60A5FA")
ws.cell(row=22, column=3, value="Audited annual reports, 12 PDFs (4 banks x 3 years)").font = Font(name="Calibri", size=10, color=WHITE)
ws.cell(row=23, column=2, value="Disclaimer:").font = Font(name="Calibri", bold=True, size=10, color="60A5FA")
ws.cell(row=23, column=3, value="For portfolio / educational purposes only. Not investment advice.").font = Font(name="Calibri", italic=True, size=9, color="64748B")
ws.cell(row=25, column=2, value="GitHub:").font = Font(name="Calibri", bold=True, size=10, color="60A5FA")
ws.cell(row=25, column=3, value="github.com/papastergiousp-maker/greek-banking-sector-analysis").font = Font(name="Calibri", size=10, color=WHITE)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — SECTOR SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════
def make_sector_summary(wb):
    ws = wb.create_sheet("Sector Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B4"

    # Column widths
    ws.column_dimensions["A"].width = 28
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 13

    # ── Title bar ─────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    c = ws.cell(row=1, column=1, value="SECTOR SUMMARY — KEY PERFORMANCE INDICATORS (FY2022–2024)")
    c.font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill      = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:M1")

    # ── Bank / year headers ───────────────────────────────────────────────────
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    # Bank group headers (merged across 3 years)
    bank_col = 2
    for bank in BANKS:
        c = ws.cell(row=2, column=bank_col, value=bank)
        c.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill      = fill(BANK_COLORS[bank])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border()
        ws.merge_cells(start_row=2, start_column=bank_col, end_row=2, end_column=bank_col+2)
        for y_idx, yr in enumerate(YEARS):
            yc = ws.cell(row=3, column=bank_col+y_idx, value=yr)
            yc.font      = Font(name="Calibri", bold=True, size=9, color=WHITE)
            yc.fill      = fill(MID_BLUE)
            yc.alignment = Alignment(horizontal="center", vertical="center")
            yc.border    = border()
        bank_col += 3

    # Metric label col header
    ws.cell(row=2, column=1).fill  = fill(DARK_BLUE)
    ws.cell(row=2, column=1).value = "Metric"
    ws.cell(row=2, column=1).font  = Font(name="Calibri", bold=True, size=10, color=WHITE)
    ws.merge_cells("A2:A3")
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="left", vertical="center")

    # ── KPI rows ──────────────────────────────────────────────────────────────
    METRICS = [
        # (label, kpi_col, fmt, best_direction, section)
        ("── Income & Profitability ──", None, None, None, "section"),
        ("NII (€m)",            "nii",            '#,##0', "high", None),
        ("Operating Income (€m)","operating_income","#,##0", "high", None),
        ("Net Profit (€m)",     "net_profit",     "#,##0", "high", None),
        ("ROE (%)",             "roe",            "0.0%", "high", None),
        ("ROA (%)",             "roa",            "0.00%", "high", None),
        ("NIM (%)",             "nim",            "0.00%", "high", None),
        ("── Efficiency ──",    None, None, None, "section"),
        ("Cost-to-Income (%)",  "cost_to_income", "0.0%", "low", None),
        ("── Balance Sheet ──", None, None, None, "section"),
        ("Loans (€m)",          "loans",          "#,##0", "high", None),
        ("Deposits (€m)",       "deposits",       "#,##0", "high", None),
        ("Total Assets (€m)",   "total_assets",   "#,##0", "high", None),
        ("Equity (€m)",         "equity",         "#,##0", "high", None),
        ("Loan-to-Deposit (%)", "loan_to_deposit","0.0%", "low", None),
        ("── Capital & Quality ──", None, None, None, "section"),
        ("CET1 (%)",            "cet1",           "0.0%", "high", None),
        ("NPE Ratio (%)",       "npe_ratio",      "0.0%", "low", None),
    ]

    data_row = 4
    for label, col_key, fmt, direction, mtype in METRICS:
        ws.row_dimensions[data_row].height = 16

        if mtype == "section":
            c = ws.cell(row=data_row, column=1, value=label)
            c.font      = Font(name="Calibri", bold=True, size=9, color=WHITE)
            c.fill      = fill(DARK_BLUE)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=13)
            data_row += 1
            continue

        # Label cell
        lc = ws.cell(row=data_row, column=1, value=label)
        lc.font      = Font(name="Calibri", size=10, color="0F172A")
        lc.fill      = fill(GREY_BG if data_row % 2 == 0 else WHITE)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border    = border()

        # Values
        bank_col = 2
        all_vals = []
        for bank in BANKS:
            for yr in YEARS:
                row = kpis[(kpis.bank == bank) & (kpis.year == yr)]
                val = float(row[col_key].values[0]) if len(row) and col_key in row.columns else None
                all_vals.append(val)
            bank_col += 3

        # Scale for percentage display
        def scale(v, fmt):
            if v is None:
                return None
            if fmt and "%" in fmt:
                return v / 100
            return v

        # Find best value
        valid_vals = [v for v in all_vals if v is not None]
        best_val = max(valid_vals) if direction == "high" else min(valid_vals)

        bank_col = 2
        val_idx  = 0
        for bank in BANKS:
            for yr in YEARS:
                raw_val = all_vals[val_idx]
                display = scale(raw_val, fmt)
                bg = GREY_BG if data_row % 2 == 0 else WHITE
                bold = False
                fg_color = "0F172A"

                # 2024 column slightly highlighted
                if yr == 2024:
                    bg = LIGHTEST

                # Best-in-period: green highlight (2024 only)
                if yr == 2024 and raw_val is not None and abs(raw_val - best_val) < 0.01:
                    bg    = LIGHT_GREEN
                    fg_color = GREEN
                    bold  = True

                vc = ws.cell(row=data_row, column=bank_col, value=display)
                vc.font      = Font(name="Calibri", size=10, bold=bold, color=fg_color)
                vc.fill      = fill(bg)
                vc.alignment = Alignment(horizontal="center", vertical="center")
                vc.border    = border()
                if fmt:
                    vc.number_format = fmt

                bank_col += 1
                val_idx  += 1

        data_row += 1

    # Freeze header rows
    ws.freeze_panes = "B4"
    return ws

make_sector_summary(wb)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEETS 3–6 — ONE PER BANK
# ═══════════════════════════════════════════════════════════════════════════════
IS_METRICS = [
    ("Net interest income",                   "Net Interest Income",        "#,##0"),
    ("Net banking fee and commission income",  "Fee & Commission Income",   "#,##0"),
    ("Operating income",                       "Operating Income",          "#,##0"),
    ("Operating expenses",                     "Operating Expenses",        "#,##0"),
    ("Profit from operations before impairments", "Pre-Provision Profit (PPOP)", "#,##0"),
    ("Impairment losses on loans",             "Impairment Losses",         "#,##0"),
    ("Profit before tax",                      "Profit Before Tax",         "#,##0"),
    ("Net profit attributable to shareholders","Net Profit",                "#,##0"),
]

BS_METRICS = [
    ("Cash and balances with central banks",   "Cash & Central Bank Balances", "#,##0"),
    ("Due from credit institutions",           "Due from Banks",               "#,##0"),
    ("Investment securities",                  "Investment Securities",         "#,##0"),
    ("Loans and advances to customers",        "Loans & Advances",             "#,##0"),
    ("Total assets",                           "Total Assets",                 "#,##0"),
    ("Due to central banks",                   "Due to Central Banks",         "#,##0"),
    ("Due to credit institutions",             "Due to Banks",                 "#,##0"),
    ("Deposits from customers",                "Customer Deposits",            "#,##0"),
    ("Total liabilities",                      "Total Liabilities",            "#,##0"),
    ("Total equity",                           "Total Equity",                 "#,##0"),
]

KPI_METRICS = [
    ("NII (€m)",            "nii",            "#,##0"),
    ("NIM (%)",             "nim",            "0.00%"),
    ("Operating Income (€m)", "operating_income", "#,##0"),
    ("Operating Expenses (€m)", "operating_expenses", "#,##0"),
    ("Net Profit (€m)",     "net_profit",     "#,##0"),
    ("ROE (%)",             "roe",            "0.0%"),
    ("ROA (%)",             "roa",            "0.00%"),
    ("Cost-to-Income (%)",  "cost_to_income", "0.0%"),
    ("Loan-to-Deposit (%)", "loan_to_deposit","0.0%"),
    ("CET1 (%)",            "cet1",           "0.0%"),
    ("NPE Ratio (%)",       "npe_ratio",      "0.0%"),
    ("Loans (€m)",          "loans",          "#,##0"),
    ("Deposits (€m)",       "deposits",       "#,##0"),
    ("Total Assets (€m)",   "total_assets",   "#,##0"),
    ("Equity (€m)",         "equity",         "#,##0"),
]

def make_bank_sheet(wb, bank):
    bcolor = BANK_COLORS[bank]
    ws = wb.create_sheet(bank)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 32
    for col in range(2, 6):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Title
    ws.row_dimensions[1].height = 30
    c = ws.cell(row=1, column=1, value=f"{bank.upper()} — FINANCIAL SUMMARY (FY2022–2024)")
    c.font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill      = fill(bcolor)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:E1")

    def section_header(row, title):
        ws.row_dimensions[row].height = 18
        c = ws.cell(row=row, column=1, value=title)
        c.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill      = fill(DARK_BLUE)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"A{row}:E{row}")
        return row + 1

    def year_headers(row):
        ws.row_dimensions[row].height = 16
        for j, yr in enumerate(YEARS, 2):
            c = ws.cell(row=row, column=j, value=yr)
            c.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
            c.fill      = fill(MID_BLUE)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = border()
        c_lbl = ws.cell(row=row, column=1, value="(€m unless stated)")
        c_lbl.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c_lbl.fill = fill(DARK_BLUE)
        c_lbl.alignment = Alignment(horizontal="left", vertical="center")
        c_lbl.border = border()
        return row + 1

    def data_rows(ws, start_row, metrics_list, source_df, is_kpi=False):
        row = start_row
        for label, col_key, fmt in metrics_list:
            ws.row_dimensions[row].height = 16
            bg = GREY_BG if row % 2 == 0 else WHITE

            # Bold totals
            is_total = any(t in label for t in ["Total", "Net Profit", "PPOP", "Pre-Provision"])
            lc = ws.cell(row=row, column=1, value=label)
            lc.font      = Font(name="Calibri", bold=is_total, size=10, color="0F172A")
            lc.fill      = fill(bg)
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            lc.border    = border()

            for j, yr in enumerate(YEARS, 2):
                if is_kpi:
                    row_data = kpis[(kpis.bank == bank) & (kpis.year == yr)]
                    val = float(row_data[col_key].values[0]) if len(row_data) and col_key in row_data.columns and pd.notna(row_data[col_key].values[0]) else None
                else:
                    row_data = source_df[(source_df.bank == bank) & (source_df.year == yr)]
                    val = float(row_data[col_key].values[0]) if len(row_data) and col_key in row_data.columns and pd.notna(row_data[col_key].values[0]) else None

                display = val / 100 if (val is not None and fmt and "%" in fmt) else val

                vc = ws.cell(row=row, column=j, value=display)
                vc.font      = Font(name="Calibri", bold=is_total, size=10, color="0F172A")
                vc.fill      = fill(LIGHTEST if yr == 2024 else bg)
                vc.alignment = Alignment(horizontal="center", vertical="center")
                vc.border    = border()
                if fmt:
                    vc.number_format = fmt

            # YoY % change in col 5 (2024 vs 2023)
            try:
                val_24 = float(kpis[(kpis.bank==bank)&(kpis.year==2024)][col_key].values[0]) if is_kpi else float(source_df[(source_df.bank==bank)&(source_df.year==2024)][col_key].values[0])
                val_23 = float(kpis[(kpis.bank==bank)&(kpis.year==2023)][col_key].values[0]) if is_kpi else float(source_df[(source_df.bank==bank)&(source_df.year==2023)][col_key].values[0])
                yoy = (val_24 - val_23) / abs(val_23) if val_23 != 0 else None
                dc = ws.cell(row=row, column=5, value=yoy)
                dc.number_format = "▲ 0.0%;▼ 0.0%;—"
                dc.font  = Font(name="Calibri", size=9,
                                color=GREEN if yoy and yoy > 0 else (RED if yoy and yoy < 0 else "64748B"))
                dc.fill  = fill(bg)
                dc.alignment = Alignment(horizontal="center", vertical="center")
                dc.border = border()
            except Exception:
                ws.cell(row=row, column=5, value="—").border = border()

            row += 1
        return row

    cur_row = 2
    cur_row = section_header(cur_row, "INCOME STATEMENT (€m)")
    cur_row = year_headers(cur_row)
    # Add YoY header
    ws.cell(row=cur_row-1, column=5, value="YoY 24/23").font = Font(name="Calibri", bold=True, size=9, color=WHITE)
    ws.cell(row=cur_row-1, column=5).fill = fill(MID_BLUE)
    ws.cell(row=cur_row-1, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=cur_row-1, column=5).border = border()
    cur_row = data_rows(ws, cur_row, IS_METRICS, income_p)

    cur_row += 1
    cur_row = section_header(cur_row, "BALANCE SHEET (€m)")
    cur_row = year_headers(cur_row)
    ws.cell(row=cur_row-1, column=5, value="YoY 24/23").font = Font(name="Calibri", bold=True, size=9, color=WHITE)
    ws.cell(row=cur_row-1, column=5).fill = fill(MID_BLUE)
    ws.cell(row=cur_row-1, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=cur_row-1, column=5).border = border()
    cur_row = data_rows(ws, cur_row, BS_METRICS, bs_p)

    cur_row += 1
    cur_row = section_header(cur_row, "KEY PERFORMANCE INDICATORS")
    cur_row = year_headers(cur_row)
    ws.cell(row=cur_row-1, column=5, value="YoY 24/23").font = Font(name="Calibri", bold=True, size=9, color=WHITE)
    ws.cell(row=cur_row-1, column=5).fill = fill(MID_BLUE)
    ws.cell(row=cur_row-1, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=cur_row-1, column=5).border = border()
    cur_row = data_rows(ws, cur_row, KPI_METRICS, None, is_kpi=True)

    ws.freeze_panes = "B4"

for bank in BANKS:
    make_bank_sheet(wb, bank)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — ASSUMPTIONS
# ═══════════════════════════════════════════════════════════════════════════════
def make_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 40

    ws.row_dimensions[1].height = 28
    c = ws.cell(row=1, column=1, value="FORECAST ASSUMPTIONS & SCENARIO DEFINITIONS")
    c.font      = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill      = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:E1")

    rows = [
        ("section", "ECB DEPOSIT FACILITY RATE SCENARIOS (%)", None, None, None, None),
        ("header",  "Scenario", "2025E", "2026E", "vs 2024 (3.15%)", "Notes"),
        ("data",    "Dovish (rapid cuts)",  1.75, 1.50, "−165bp / −165bp", "Faster normalisation; inflation below target"),
        ("data",    "Base (gradual cuts)",  2.25, 2.00, "−90bp / −115bp",  "ECB consensus path as of early 2025"),
        ("data",    "Hawkish (cuts pause)", 3.15, 3.15, "0bp / 0bp",       "Inflation re-acceleration / geopolitical shock"),
        (None, None, None, None, None, None),
        ("section", "NII SENSITIVITY", None, None, None, None),
        ("header",  "Parameter", "Value", "Unit", "", "Source"),
        ("data",    "Sector NII impact per 25bp ECB cut", -195, "€m",  "", "Analyst estimate; midpoint of €170–220m range"),
        ("data",    "Allocation basis",    "Proportional to 2024 NII share", "", "", "Applied per bank in forecast model"),
        ("data",    "Loan volume growth (base)", 2.0, "% p.a.", "", "Assumed for 2025E and 2026E"),
        (None, None, None, None, None, None),
        ("section", "COST OF EQUITY (CoE) — GORDON GROWTH P/B", None, None, None, None),
        ("header",  "Component", "Value", "Unit", "", "Notes"),
        ("data",    "Risk-free rate (Rf)",      3.5,  "%", "", "German 10Y Bund approximate"),
        ("data",    "Equity Risk Premium (ERP)", 5.5, "%", "", "European market estimate"),
        ("data",    "Country Risk Premium (CRP)", 2.0, "%", "", "Greece sovereign spread adjustment"),
        ("data",    "Implied CoE (Rf + β×ERP + CRP)", 11.0, "%", "", "β assumed ≈ 1.0 for Greek banks"),
        (None, None, None, None, None, None),
        ("section", "STRESS TEST PARAMETERS (EBA-STYLE ADVERSE)", None, None, None, None),
        ("header",  "Shock", "Value", "Unit", "", "Notes"),
        ("data",    "Additional Cost of Risk",  200, "bps", "", "Applied to stressed loan book"),
        ("data",    "Loan Volume Decline",      -15, "%",  "", "Gross loan book reduction"),
        ("data",    "NIM Compression",          -50, "bps", "", "Applied to total assets"),
        ("data",    "RWA Proxy",                50,  "% of loans", "", "Loan book risk-weighted at 50%"),
        ("data",    "Tax Rate",                 "Implied", "%", "", "From each bank's FY2024 actuals (capped 15–35%)"),
        (None, None, None, None, None, None),
        ("section", "DATA QUALITY", None, None, None, None),
        ("header",  "Check", "Result", "", "", ""),
        ("data",    "pytest suite",             "26/26 passed", "", "", "ROE, ROA, C/I, L/D, NIM recomputed from CSVs"),
        ("data",    "Balance sheet identity",   "Within ±1%", "", "", "Assets ≈ Liabilities + Equity for all 12 bank-years"),
        ("data",    "NPE ratios verified",      "12/12 from PDF", "", "", "Each figure cited to source PDF page"),
    ]

    section_fill = fill(DARK_BLUE)
    header_fill  = fill(MID_BLUE)
    data_fills   = [fill(WHITE), fill(GREY_BG)]

    data_count = 0
    for idx, row_data in enumerate(rows, 2):
        rtype = row_data[0]
        if rtype is None:
            ws.row_dimensions[idx].height = 8
            continue
        ws.row_dimensions[idx].height = 16
        if rtype == "section":
            c = ws.cell(row=idx, column=1, value=row_data[1])
            c.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
            c.fill      = section_fill
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.merge_cells(f"A{idx}:E{idx}")
        elif rtype == "header":
            for j, val in enumerate(row_data[1:], 1):
                c = ws.cell(row=idx, column=j, value=val)
                c.font      = Font(name="Calibri", bold=True, size=9, color=WHITE)
                c.fill      = header_fill
                c.alignment = Alignment(horizontal="center" if j>1 else "left", vertical="center")
                c.border    = border()
        elif rtype == "data":
            bg = data_fills[data_count % 2]
            data_count += 1
            for j, val in enumerate(row_data[1:], 1):
                c = ws.cell(row=idx, column=j, value=val)
                c.font      = Font(name="Calibri", size=10, color="0F172A")
                c.fill      = bg
                c.alignment = Alignment(horizontal="center" if j>1 else "left",
                                        vertical="center", indent=1 if j==1 else 0)
                c.border    = border()

make_assumptions(wb)

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Written: {OUT}")
